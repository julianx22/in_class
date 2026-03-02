from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth import views as auth_views
from django.shortcuts import render, redirect
from django.urls import reverse, reverse_lazy
import math

from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.cache import never_cache
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db import connection

from datetime import datetime, timedelta, date

from .forms import ProfileEditForm, PublicRegistrationForm
from .decorators import (
    login_required_any,
    instructor_required,
    aprendiz_required,
)

from inclass_legacy.sync import upsert_usuario_from_django

import os
import uuid


# =========================
# Configuración de asistencia
# =========================

TOLERANCIA_MINUTOS = 25  # minutos de tolerancia para marcar 'Presente'


# =========================
# helpers internos
# =========================

def _resolve_username_maybe_email(username_or_email: str) -> str:
    """
    Si el usuario escribe un correo, buscamos en auth_user por email (case-insensitive).
    Si hay varios con el mismo correo, tomamos:
      - primero un superusuario (si existe),
      - si no, el primero por id.
    Si no encuentra nada, devolvemos lo mismo que entró.
    """
    if '@' in username_or_email:
        User = get_user_model()
        qs = User.objects.filter(email__iexact=username_or_email)

        if not qs.exists():
            # No hay nadie con ese correo, intentamos autenticar tal cual
            return username_or_email

        # Preferir superuser si lo hay, si no, el primero por id
        u = qs.order_by('-is_superuser', 'id').first()
        return u.get_username()

    # Si no contiene @ asumimos que ya es username
    return username_or_email

def _user_in_group(user, name: str) -> bool:
    """
    name debe ser el nombre EXACTO del grupo en Django: 'aprendiz', 'instructor', 'admin', etc.
    """
    return user.is_authenticated and user.groups.filter(name=name).exists()


def _save_uploaded_file(user_id: int, fobj):
    """
    Guarda un archivo subido en MEDIA_ROOT/justificaciones/<user_id>/
    y retorna la ruta relativa (para guardar en DB).
    """
    if not fobj:
        return None
    base_dir = os.path.join(settings.MEDIA_ROOT, "justificaciones", str(user_id))
    os.makedirs(base_dir, exist_ok=True)
    ext = os.path.splitext(fobj.name)[1]
    safe_name = timezone.localtime().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:6] + (ext or "")
    final_path = os.path.join(base_dir, safe_name)
    with open(final_path, "wb+") as dest:
        for chunk in fobj.chunks():
            dest.write(chunk)
    rel = os.path.relpath(final_path, settings.MEDIA_ROOT).replace("\\", "/")
    return rel


# =========================
# auth views
# =========================

@never_cache
def login_view(request):
    # Si ya está autenticado, no mostramos el login ni reprocesamos el POST
    if request.user.is_authenticated:
        if request.user.is_staff or request.user.is_superuser:
            return redirect(reverse("admin:index"))
        return redirect("core:home")

    if request.method == "POST":
        raw_username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        username = _resolve_username_maybe_email(raw_username)

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)

            if user.is_staff or user.is_superuser:
                return redirect(reverse("admin:index"))

            next_url = request.POST.get("next") or request.GET.get("next")
            if next_url:
                return redirect(next_url)

            return redirect("core:home")
        else:
            messages.error(request, "Credenciales inválidas.")
    return render(request, "registration/login.html")


@never_cache
def register_view(request):
    """
    Registro público de usuario:
    - Crea un usuario en auth_user
    - Crea/actualiza el registro en dbo.Usuario como aprendiz genérico
    - Luego el admin asigna rol, ficha y programa en el admin de Django.
    """
    if request.user.is_authenticated:
        return redirect("core:home")

    if request.method == "POST":
        form = PublicRegistrationForm(request.POST)
        if form.is_valid():
            # 1) Crear usuario Django
            user = form.save(commit=False)
            user.is_active = True
            user.save()

            # 2) Crear/actualizar en dbo.Usuario como aprendiz sin programa/ficha
            try:
                upsert_usuario_from_django(
                    user,
                    role="aprendiz",
                    programa=None,
                    jornada="",
                    ficha="",
                )
            except Exception:
                # No rompemos el registro web si falla SQL Server
                pass

            # 3) Loguear con backend explícito (para evitar conflicto de backends)
            login(
                request,
                user,
                backend="django.contrib.auth.backends.ModelBackend",
            )

            messages.success(
                request,
                "Tu cuenta ha sido creada. Un administrador te asignará rol, ficha y programa."
            )
            return redirect("core:home")
        else:
            messages.error(request, "Revisa los campos del formulario.")
    else:
        form = PublicRegistrationForm()

    return render(request, "registration/register.html", {"form": form})


def logout_view(request):
    logout(request)
    return redirect(settings.LOGOUT_REDIRECT_URL)


class CustomPasswordResetView(auth_views.PasswordResetView):
    """
    Vista de 'Olvidaste tu contraseña' que valida si el correo existe
    en la tabla de usuarios. Si no existe, muestra un error en el formulario.
    """
    template_name = "registration/password_reset_form.html"
    email_template_name = "registration/password_reset_email.txt"
    subject_template_name = "registration/password_reset_subject.txt"
    success_url = reverse_lazy("core:password_reset_done")

    def form_valid(self, form):
        email = (form.cleaned_data.get("email") or "").strip()
        User = get_user_model()
        existe = User.objects.filter(email__iexact=email, is_active=True).exists()

        if not existe:
            form.add_error(
                None,
                "No existe ningún usuario registrado con este correo."
            )
            return self.form_invalid(form)

        return super().form_valid(form)


# =========================
# home inteligente
# =========================

@login_required_any
@never_cache
def home_redirect(request):
    """
    Redirección inteligente según rol.
    - Instructor  → dash_instructor
    - Aprendiz    → dash_aprendiz
    - Sin rol     → mensaje de “cuenta pendiente” (sin bucles)
    """
    u = request.user

    # 1) Instructores / admin
    if _user_in_group(u, "instructor") or u.is_staff:
        return redirect("core:dash_instructor")

    # 2) Aprendices
    if _user_in_group(u, "aprendiz"):
        return redirect("core:dash_aprendiz")

    # 3) Usuario autenticado pero SIN grupo/rol todavía
    return HttpResponse(
        """
        <html>
          <head>
            <meta charset="utf-8">
            <title>Cuenta pendiente</title>
          </head>
          <body style="font-family:system-ui, -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
                       background:#020617; color:#e5e7eb; display:flex; align-items:center;
                       justify-content:center; height:100vh; margin:0;">
            <div style="max-width:520px; padding:24px; border-radius:16px;
                        border:1px solid #1f2937; background:#020617;">
              <h1 style="margin-top:0; font-size:20px; margin-bottom:8px;">
                Tu cuenta está casi lista 👀
              </h1>
              <p style="font-size:14px; line-height:1.5; margin:0 0 8px 0;">
                Ya creaste tu usuario y estás autenticado, pero todavía no tienes
                <strong>rol, ficha ni programa</strong> asignados.
              </p>
              <p style="font-size:14px; line-height:1.5; margin:0 0 8px 0;">
                Un administrador debe entrar al panel de administración, abrir tu usuario
                y elegir el rol (<strong>Aprendiz / Instructor / Admin</strong>) y el programa.
              </p>
              <p style="font-size:14px; line-height:1.5; margin:0;">
                Cuando el admin guarde esos datos, podrás entrar normalmente a tu panel
                de asistencia desde esta misma cuenta.
              </p>
            </div>
          </body>
        </html>
        """,
    )


# =========================
# dashboards
# =========================

@aprendiz_required
@never_cache
def dash_aprendiz(request):
    return render(request, "dash/aprendiz.html")


@instructor_required
@never_cache
def dash_instructor(request):
    """
    Panel del instructor + horario desde la tabla Horario_Instructor.
    """
    user = request.user

    # Mapear auth_user -> Usuario.ID_Usuario
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 u.ID_Usuario
            FROM Usuario u
            JOIN auth_user au ON LOWER(au.email) = LOWER(u.Correo)
            WHERE au.id = %s
        """, [user.id])
        row = cur.fetchone()

    horario = []
    if row:
        id_instructor = row[0]
        with connection.cursor() as cur:
            cur.execute("""
                SELECT DiaSemana, HoraInicio, HoraFin, Competencia, Ficha
                FROM Horario_Instructor
                WHERE ID_Instructor = %s
                ORDER BY
                  CASE DiaSemana
                    WHEN 'Lunes' THEN 1
                    WHEN 'Martes' THEN 2
                    WHEN 'Miercoles' THEN 3
                    WHEN 'Jueves' THEN 4
                    WHEN 'Viernes' THEN 5
                    ELSE 6
                  END,
                  HoraInicio
            """, [id_instructor])
            for dia, hi, hf, comp, ficha in cur.fetchall():
                horario.append({
                    "dia": dia,
                    "hora_inicio": hi.strftime("%H:%M"),
                    "hora_fin": hf.strftime("%H:%M"),
                    "competencia": comp or "",
                    "ficha": ficha or "",
                })

    return render(request, "dash/instructor.html", {"horario": horario})


# =========================
# perfil
# =========================

@login_required_any
@never_cache
def profile_edit(request):
    if request.method == "POST":
        form = ProfileEditForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Perfil actualizado correctamente.")
            return redirect("core:profile_edit")
        else:
            messages.error(request, "Revisa los campos.")
    else:
        form = ProfileEditForm(instance=request.user)

    return render(request, "profile/edit.html", {"form": form})


# =========================
# QR (pantalla + APIs)
# =========================

@instructor_required
@never_cache
def qr_generar(request):
    return render(request, "dash/qr_generar.html")


@instructor_required
@require_POST
def qr_generar_api(request):
    """
    JSON esperado:
    {
      "ficha": "...",
      "programa": "...",
      "token": "QR-ABC123",
      "duracion": 2
    }

    La jornada no viene del formulario:
    - Si ya existe (programa + ficha) en Programa, se usa esa jornada.
    - Si no existe, se crea el registro copiando la jornada de otro programa con mismo nombre,
      y si tampoco hay, se usa 'Sin jornada'.
    """
    import json

    body_raw = request.body.decode("utf-8") or "{}"
    body = json.loads(body_raw)

    ficha = (body.get("ficha") or "").strip()
    programa = (body.get("programa") or "").strip()
    token = (body.get("token") or "").strip()

    try:
        duracion = int(body.get("duracion") or 0)
    except Exception:
        duracion = 0

    user_id = request.user.id

    if not ficha or not programa or not token or duracion <= 0:
        return JsonResponse({"ok": False, "msg": "Datos incompletos."}, status=400)

    # 1) buscar/crear Programa (Ficha + NombrePrograma)
    with connection.cursor() as cur:
        cur.execute("""
            SELECT ID_Programa, Jornada
            FROM Programa
            WHERE Ficha = %s
              AND NombrePrograma = %s
        """, [ficha, programa])
        row = cur.fetchone()

    if row:
        id_programa = row[0]
        jornada = row[1] or ""
    else:
        # Intentar copiar jornada de otro registro del mismo programa
        with connection.cursor() as cur:
            cur.execute("""
                SELECT TOP 1 Jornada
                FROM Programa
                WHERE NombrePrograma = %s
                ORDER BY ID_Programa DESC
            """, [programa])
            jrow = cur.fetchone()

        jornada = (jrow[0] if jrow else "") or "Sin jornada"

        with connection.cursor() as cur:
            cur.execute("""
                INSERT INTO Programa (NombrePrograma, Jornada, Ficha)
                OUTPUT INSERTED.ID_Programa
                VALUES (%s, %s, %s)
            """, [programa, jornada, ficha])
            new_row = cur.fetchone()

        if not new_row:
            return JsonResponse({"ok": False, "msg": "No pude crear el programa."}, status=500)

        id_programa = new_row[0]

    # 2) insertar QR (con hora LOCAL y DuracionMin)
    now = timezone.localtime()
    fecha_sql = now.date()
    hora_sql = now.strftime("%H:%M:%S")

    with connection.cursor() as cur:
        cur.execute("""
            INSERT INTO Codigo_Generado (TipoCodigo, Fecha, Codigo, Hora, ID_Usuario, ID_Programa, DuracionMin)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, ["QR", fecha_sql, token, hora_sql, user_id, id_programa, duracion])

    # 3) NOTIFICACIONES: avisar a aprendices activos de ese programa
    titulo = "QR activo"
    cuerpo = f"Código {token} activo por {duracion} min. Programa: {programa} • Ficha {ficha} • {jornada}"
    with connection.cursor() as cur:
        cur.execute("""
            INSERT INTO Notificacion (ID_Usuario, Titulo, Cuerpo, Tipo)
            SELECT au.id, %s, %s, 'alerta'
            FROM Usuario u
            JOIN auth_user au ON LOWER(au.email) = LOWER(u.Correo)
            WHERE u.Estado = 'Activo' AND u.ID_Programa = %s
        """, [titulo, cuerpo, id_programa])

    return JsonResponse({"ok": True, "msg": "Código guardado.", "programa_id": id_programa})


@instructor_required
@require_GET
def api_fichas_por_programa(request):
    """
    Devuelve las fichas registradas para un programa en la tabla Programa.
    """
    programa = (request.GET.get("programa") or "").strip()
    if not programa:
        return JsonResponse({"ok": False, "msg": "Falta 'programa'."}, status=400)

    with connection.cursor() as cur:
        cur.execute("""
            SELECT Ficha, Jornada
            FROM Programa
            WHERE NombrePrograma = %s
            ORDER BY Ficha
        """, [programa])
        rows = cur.fetchall()

    out = []
    for ficha, jornada in rows:
        out.append({
            "ficha": ficha or "",
            "jornada": jornada or "",
        })

    return JsonResponse({"ok": True, "rows": out})


@instructor_required
@require_GET
def api_programa_por_ficha(request):
    ficha = (request.GET.get("ficha") or "").strip()
    if not ficha:
        return JsonResponse({"ok": False, "msg": "Falta 'ficha'."}, status=400)

    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 ID_Programa, NombrePrograma, Jornada
            FROM Programa
            WHERE Ficha = %s
            ORDER BY ID_Programa DESC
        """, [ficha])
        row = cur.fetchone()

    if not row:
        return JsonResponse({"ok": False, "msg": "No encontrado."}, status=404)

    id_programa, nombre_programa, jornada = row
    return JsonResponse({"ok": True, "id_Programa": id_programa, "programa": nombre_programa, "jornada": jornada})


# =========================
# Registrar asistencia (pantalla + API)
# =========================

@aprendiz_required
@never_cache
def registrar_asistencia(request):
    return render(request, "dash/registrar_asistencia.html")


@aprendiz_required
@require_POST
def registrar_asistencia_api(request):
    """
    Recibe { token: "QR-ABC123" } y registra 'Entrada' con Estado:
      - 'Presente' si está dentro de la tolerancia (TOLERANCIA_MINUTOS)
      - 'Tarde' si está después de la tolerancia pero antes de que el QR expire
    Valida expiración y que el aprendiz pertenezca al mismo programa/ficha del QR.
    """
    import json
    body = json.loads(request.body.decode("utf-8"))
    token = (body.get("token") or "").strip()
    if not token:
        return JsonResponse({"ok": False, "msg": "Falta el token."}, status=400)

    # 1) QR
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1
                cg.ID_Codigo, cg.ID_Programa, cg.Fecha, cg.Hora, cg.DuracionMin,
                p.NombrePrograma
            FROM Codigo_Generado cg
            JOIN Programa p ON p.ID_Programa = cg.ID_Programa
            WHERE cg.Codigo = %s
            ORDER BY cg.ID_Codigo DESC
        """, [token])
        row = cur.fetchone()

    if not row:
        return JsonResponse({"ok": False, "msg": "Código no encontrado."}, status=404)

    id_codigo, id_programa, fecha, hora, dur_min, nombre_programa = row

    # 2) Validar que el aprendiz pertenece a ese programa (map Usuario por correo)
    with connection.cursor() as cur:
        cur.execute("""
            SELECT u.ID_Programa
            FROM Usuario u
            JOIN auth_user au ON LOWER(au.email) = LOWER(u.Correo)
            WHERE au.id = %s
        """, [request.user.id])
        prog_row = cur.fetchone()

    if not prog_row:
        return JsonResponse({"ok": False, "msg": "Tu usuario no está asociado a ningún programa."}, status=403)

    user_programa = prog_row[0]
    if user_programa != id_programa:
        return JsonResponse({"ok": False, "msg": "Este código QR no corresponde a tu ficha/programa."}, status=403)

    # 3) Expiración + cálculo de Presente/Tarde
    tz = timezone.get_current_timezone()
    created_naive = datetime.combine(fecha, hora)
    created = timezone.make_aware(created_naive, tz)
    now = timezone.localtime()
    expires = created + timedelta(minutes=int(dur_min or 0))

    if now >= expires:
        # QR totalmente vencido: NO se registra nada en Asistencia
        return JsonResponse({"ok": False, "msg": "El código ya expiró."}, status=410)

    # minutos desde que se generó el QR
    delta_min = max((now - created).total_seconds() / 60.0, 0)

    if delta_min <= TOLERANCIA_MINUTOS:
        estado = "Presente"
    else:
        estado = "Tarde"

    # 4) Insert asistencia
    fecha_sql = now.date()
    hora_sql = now.strftime("%H:%M:%S")

    with connection.cursor() as cur:
        cur.execute("""
            INSERT INTO Asistencia (TipoRegistro, FechaRegistro, HoraRegistro, Estado, ID_Usuario, ID_Programa, ID_Codigo)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, ["Entrada", fecha_sql, hora_sql, estado, request.user.id, id_programa, id_codigo])

    # 5) Notificación
    with connection.cursor() as cur:
        cur.execute("""
            INSERT INTO Notificacion (ID_Usuario, Titulo, Cuerpo, Tipo)
            VALUES (%s, %s, %s, %s)
        """, [request.user.id, "Asistencia registrada", f"{nombre_programa} • {hora_sql}", "info"])

    return JsonResponse({
        "ok": True,
        "msg": f"Asistencia registrada ({estado}).",
        "programa": nombre_programa,
        "hora": hora_sql,
        "estado": estado,
    })


# =========================
# Router QR + Auto-registro + Live list
# =========================

def qr_scan_router(request, code: str):
    request.session["pending_qr_code"] = code

    if not request.user.is_authenticated:
        return redirect(f"{reverse('core:login')}?next={request.path}")

    if _user_in_group(request.user, "aprendiz") or (
        not _user_in_group(request.user, "instructor") and not request.user.is_staff
    ):
        return redirect("core:registrar_asistencia_auto", code=code)

    messages.info(request, "Ese QR es para aprendices.")
    return redirect("core:dash_instructor")


@aprendiz_required
@never_cache
def registrar_asistencia_auto(request, code: str):
    token = (code or "").strip()
    if not token:
        return HttpResponse("<h2>QR inválido</h2>", status=400)

    # 1) QR
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 cg.ID_Codigo, cg.ID_Programa, p.NombrePrograma,
                           cg.Fecha, cg.Hora, cg.DuracionMin
            FROM Codigo_Generado cg
            JOIN Programa p ON p.ID_Programa = cg.ID_Programa
            WHERE cg.Codigo = %s
            ORDER BY cg.ID_Codigo DESC
        """, [token])
        row = cur.fetchone()

    if not row:
        return HttpResponse("<h2>Código no encontrado o expirado.</h2>", status=404)

    id_codigo, id_programa, nombre_programa, fecha, hora, dur_min = row

    # 2) Validar que el aprendiz pertenece a ese programa
    with connection.cursor() as cur:
        cur.execute("""
            SELECT u.ID_Programa
            FROM Usuario u
            JOIN auth_user au ON LOWER(au.email) = LOWER(u.Correo)
            WHERE au.id = %s
        """, [request.user.id])
        prog_row = cur.fetchone()

    if not prog_row:
        return HttpResponse("<h2>No estás asociado a ningún programa.</h2>", status=403)

    user_programa = prog_row[0]
    if user_programa != id_programa:
        return HttpResponse("<h2>Este código QR no corresponde a tu ficha/programa.</h2>", status=403)

    # 3) Expiración + cálculo de Presente/Tarde
    tz = timezone.get_current_timezone()
    created = timezone.make_aware(datetime.combine(fecha, hora), tz)
    now = timezone.localtime()
    expires = created + timedelta(minutes=int(dur_min or 0))

    if now >= expires:
        return HttpResponse("<h2>Código expirado.</h2>", status=410)

    delta_min = max((now - created).total_seconds() / 60.0, 0)

    if delta_min <= TOLERANCIA_MINUTOS:
        estado = "Presente"
    else:
        estado = "Tarde"

    fecha_sql = now.date()
    hora_sql = now.strftime("%H:%M:%S")

    # 4) Insert asistencia
    with connection.cursor() as cur:
        cur.execute("""
            INSERT INTO Asistencia (TipoRegistro, FechaRegistro, HoraRegistro, Estado, ID_Usuario, ID_Programa, ID_Codigo)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, ["Entrada", fecha_sql, hora_sql, estado, request.user.id, id_programa, id_codigo])

    # 5) Notificación
    with connection.cursor() as cur:
        cur.execute("""
            INSERT INTO Notificacion (ID_Usuario, Titulo, Cuerpo, Tipo)
            VALUES (%s, %s, %s, %s)
        """, [request.user.id, "Asistencia registrada", f"{nombre_programa} • {hora_sql}", "info"])

    html = f"""
    <html>
      <head>
        <meta name="viewport" content="width=device-width, initial-scale=1"/>
        <title>Asistencia registrada</title>
        <style>
          body {{ background:#0b0f14;color:#e5e7eb;font-family:ui-sans-serif,system-ui;display:grid;place-items:center;height:100vh; }}
          .box {{ background:rgba(15,23,42,.6); padding:20px 24px; border-radius:14px; border:1px solid rgba(148,163,184,.22); text-align:center; }}
          h1{{ margin:0 0 8px; font-size:18px }}
          p {{ margin:6px 0; font-size:14px; color:#9aa6b2 }}
          a.btn {{ display:inline-block; margin-top:10px; padding:10px 14px; border-radius:10px;
                   background:linear-gradient(90deg,#10b981,#22d3ee); color:#0f172a; font-weight:800; text-decoration:none; }}
        </style>
      </head>
      <body>
        <div class="box">
          <h1>✅ Asistencia registrada ({estado})</h1>
          <p>{nombre_programa}</p>
          <p>Hora: {hora_sql}</p>
          <a class="btn" href="{reverse('core:dash_aprendiz')}">Ir al panel</a>
        </div>
      </body>
    </html>
    """
    return HttpResponse(html)


@instructor_required
def asistencia_list_api(request, code: str):
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 ID_Codigo
            FROM Codigo_Generado
            WHERE Codigo = %s
            ORDER BY ID_Codigo DESC
        """, [code])
        row = cur.fetchone()

    if not row:
        return JsonResponse({"rows": []})

    id_codigo = row[0]

    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 20
                a.FechaRegistro, a.HoraRegistro, a.Estado,
                COALESCE(NULLIF(LTRIM(RTRIM(u.first_name + ' ' + u.last_name)), ''), u.username) AS Nombre
            FROM Asistencia a
            LEFT JOIN auth_user u ON u.id = a.ID_Usuario
            WHERE a.ID_Codigo = %s
            ORDER BY a.FechaRegistro DESC, a.HoraRegistro DESC
        """, [id_codigo])
        data = cur.fetchall()

    out = []
    for f, h, estado, nombre in data:
        fecha_txt = f"{f} {h}"
        out.append({"fecha": fecha_txt, "estado": estado, "nombre": nombre or ""})

    return JsonResponse({"rows": out})


# ===== QR activo para el aprendiz (último no expirado) =====

@require_GET
@aprendiz_required
def api_qr_activo(request):
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1
                cg.Codigo, cg.Fecha, cg.Hora, cg.DuracionMin,
                p.NombrePrograma, p.Ficha, p.Jornada
            FROM Codigo_Generado cg
            JOIN Programa p ON p.ID_Programa = cg.ID_Programa
            ORDER BY cg.ID_Codigo DESC
        """)
        row = cur.fetchone()

    if not row:
        return JsonResponse({"ok": False})

    codigo, fecha, hora, dur_min, nombre_programa, ficha, jornada = row
    if not dur_min or dur_min <= 0:
        return JsonResponse({"ok": False})

    tz = timezone.get_current_timezone()
    created = timezone.make_aware(datetime.combine(fecha, hora), tz)
    now = timezone.localtime()
    expires = created + timedelta(minutes=dur_min)

    if now >= expires:
        return JsonResponse({"ok": False})

    seconds_left = int((expires - now).total_seconds())
    return JsonResponse({
        "ok": True,
        "code": codigo,
        "seconds_left": seconds_left,
        "programa": nombre_programa,
        "ficha": ficha,
        "jornada": jornada,
        "expires_at": expires.isoformat(),
    })


# =========================
# páginas públicas
# =========================

def acerca_view(request):
    return render(request, "pages/acerca.html")


def ayuda_view(request):
    return render(request, "pages/ayuda.html")


def contacto_view(request):
    return render(request, "pages/contacto.html")


def creditos_view(request):
    return render(request, "pages/creditos.html")


# ============================================================
# ================== NOTIFICACIONES (NUEVO) ==================
# ============================================================

@aprendiz_required
@never_cache
def notificaciones_view(request):
    """Pantalla de notificaciones del aprendiz."""
    return render(request, "dash/notificaciones.html")


@aprendiz_required
@require_GET
def notificaciones_list_api(request):
    uid = request.user.id
    try:
        page = max(int(request.GET.get("page", 1)), 1)
    except Exception:
        page = 1
    try:
        page_size = int(request.GET.get("page_size", 8))
        page_size = max(1, min(page_size, 50))
    except Exception:
        page_size = 8

    only_unread = str(request.GET.get("only_unread", "0")).lower() in ("1", "true", "yes")

    where = "WHERE ID_Usuario = %s" + (" AND Leida = 0" if only_unread else "")
    params = [uid]

    # Conteo total
    with connection.cursor() as cur:
        cur.execute(f"SELECT COUNT(*) FROM Notificacion {where}", params)
        total_row = cur.fetchone()
        total = total_row[0] if total_row else 0

    pages = max((total + page_size - 1) // page_size, 1)
    if page > pages:
        page = pages
    offset = (page - 1) * page_size

    rows = []
    # Paginado
    with connection.cursor() as cur:
        cur.execute(f"""
            SELECT ID_Notificacion, Titulo, Cuerpo, Tipo, Leida, CreadoEn
            FROM Notificacion
            {where}
            ORDER BY CreadoEn DESC, ID_Notificacion DESC
            OFFSET %s ROWS FETCH NEXT %s ROWS ONLY
        """, params + [offset, page_size])
        for (nid, titulo, cuerpo, tipo, leida, creado) in cur.fetchall():
            rows.append({
                "id": nid,
                "titulo": titulo or "",
                "cuerpo": cuerpo or "",
                "tipo": (tipo or "info").lower(),
                "leida": bool(leida),
                "creado_en": creado.isoformat() if creado else ""
            })

    # Conteo de no leídas
    with connection.cursor() as cur:
        cur.execute(
            "SELECT COUNT(*) FROM Notificacion WHERE ID_Usuario=%s AND Leida=0",
            [uid],
        )
        unread_row = cur.fetchone()
        unread = unread_row[0] if unread_row else 0

    return JsonResponse({
        "rows": rows,
        "page": page,
        "pages": pages,
        "total": total,
        "unread": unread,
    })


@aprendiz_required
@require_POST
def notificacion_mark_read_api(request):
    import json
    body = json.loads(request.body.decode("utf-8") or "{}")
    try:
        nid = int(body.get("id") or 0)
    except Exception:
        nid = 0
    if nid <= 0:
        return JsonResponse({"ok": False, "msg": "ID inválido."}, status=400)

    with connection.cursor() as cur:
        cur.execute("""
            UPDATE Notificacion
               SET Leida = 1
             WHERE ID_Notificacion = %s AND ID_Usuario = %s
        """, [nid, request.user.id])
        updated = cur.rowcount > 0

    return JsonResponse({"ok": updated})


@aprendiz_required
@require_POST
def notificacion_mark_all_api(request):
    with connection.cursor() as cur:
        cur.execute("""
            UPDATE Notificacion SET Leida = 1
            WHERE ID_Usuario = %s AND Leida = 0
        """, [request.user.id])
    return JsonResponse({"ok": True})


@aprendiz_required
@require_GET
def notif_unread_count_api(request):
    """
    Devuelve solo el número de notificaciones no leídas
    para el usuario actual.
    """
    with connection.cursor() as cur:
        cur.execute(
            "SELECT COUNT(*) FROM Notificacion WHERE ID_Usuario=%s AND Leida=0",
            [request.user.id],
        )
        row = cur.fetchone()

    unread = row[0] if row else 0
    return JsonResponse({"count": unread, "unread": unread})


# ============================================================
# ================= HISTORIAL DE ASISTENCIA ==================
# ============================================================

@aprendiz_required
@never_cache
def historial_asistencia_view(request):
    """Pantalla del historial del aprendiz."""
    return render(request, "dash/historial_asistencia.html")


@aprendiz_required
@require_GET
def historial_asistencia_api(request):
    """
    Devuelve el historial de asistencia del aprendiz logueado.

    Soporta:
      - Filtros por fecha (?desde=YYYY-MM-DD&hasta=YYYY-MM-DD)
      - Filtro por estado (?estado=Presente|Tarde|Ausente|Justificado|todos)
      - Paginación (?page, ?page_size)
      - Estadísticas para las tarjetas de resumen.
    """
    uid = request.user.id

    # Filtros de fecha y estado
    desde_s = (request.GET.get("desde") or "").strip()
    hasta_s = (request.GET.get("hasta") or "").strip()
    estado = (request.GET.get("estado") or "").strip()

    desde = None
    hasta = None

    if desde_s:
        try:
            desde = datetime.strptime(desde_s, "%Y-%m-%d").date()
        except Exception:
            desde = None

    if hasta_s:
        try:
            hasta = datetime.strptime(hasta_s, "%Y-%m-%d").date()
        except Exception:
            hasta = None

    try:
        page = max(int(request.GET.get("page", "1")), 1)
    except Exception:
        page = 1
    try:
        page_size = int(request.GET.get("page_size", "12"))
        page_size = max(1, min(page_size, 50))
    except Exception:
        page_size = 12

    where_clauses = ["a.ID_Usuario = %s"]
    params = [uid]

    if desde and hasta:
        where_clauses.append("a.FechaRegistro BETWEEN %s AND %s")
        params.extend([desde, hasta])
    elif desde:
        where_clauses.append("a.FechaRegistro >= %s")
        params.append(desde)
    elif hasta:
        where_clauses.append("a.FechaRegistro <= %s")
        params.append(hasta)

    if estado and estado.lower() != "todos":
        where_clauses.append("a.Estado = %s")
        params.append(estado)

    where_sql = "WHERE " + " AND ".join(where_clauses)

    # Conteo total + estadísticas para las tarjetas
    with connection.cursor() as cur:
        cur.execute(f"""
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN a.Estado IN ('Presente','Justificado') THEN 1 ELSE 0 END) AS presentes,
                SUM(CASE WHEN a.Estado = 'Tarde' THEN 1 ELSE 0 END) AS tardes,
                SUM(CASE WHEN a.Estado = 'Ausente' THEN 1 ELSE 0 END) AS ausentes
            FROM Asistencia a
            {where_sql}
        """, params)
        row = cur.fetchone()

    total = row[0] if row and row[0] is not None else 0
    presentes = row[1] if row and row[1] is not None else 0
    tardes = row[2] if row and row[2] is not None else 0
    ausentes = row[3] if row and row[3] is not None else 0

    pages = max((total + page_size - 1) // page_size, 1)
    if page > pages:
        page = pages
    offset = (page - 1) * page_size

    rows = []
    if total > 0:
        with connection.cursor() as cur:
            cur.execute(f"""
                SELECT
                    a.ID_Asistencia,
                    a.TipoRegistro,
                    a.FechaRegistro,
                    a.HoraRegistro,
                    a.Estado,
                    p.NombrePrograma,
                    p.Ficha,
                    p.Jornada
                FROM Asistencia a
                JOIN Programa p ON p.ID_Programa = a.ID_Programa
                {where_sql}
                ORDER BY a.FechaRegistro DESC, a.HoraRegistro DESC, a.ID_Asistencia DESC
                OFFSET %s ROWS FETCH NEXT %s ROWS ONLY
            """, params + [offset, page_size])
            for (ida, tipo, f, h, estado_row, prog, ficha, jornada) in cur.fetchall():
                rows.append({
                    "id": ida,
                    "tipo": tipo or "",
                    "fecha": f.isoformat() if f else "",
                    "hora": h.strftime("%H:%M:%S") if h else "",
                    "estado": estado_row or "",
                    "programa": prog or "",
                    "ficha": ficha or "",
                    "jornada": jornada or "",
                })

    return JsonResponse({
        "rows": rows,
        "page": page,
        "pages": pages,
        "total": total,
        "stats": {
            "total": total,
            "presentes": presentes,
            "tardes": tardes,
            "ausentes": ausentes,
        }
    })


# ============================================================
# ================== JUSTIFICACIONES (NUEVO) =================
# ============================================================

# ---------- Aprendiz ----------

@aprendiz_required
@never_cache
def justificaciones_aprendiz(request):
    return render(request, "dash/justificaciones.html")


@aprendiz_required
@require_GET
def justif_instructores_aprendiz_api(request):
    """
    Devuelve los instructores asignados al aprendiz actual,
    según la tabla Instructor_Asignado.
    """
    user = request.user

    # Mapear auth_user.id -> Usuario.ID_Usuario (aprendiz)
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 u.ID_Usuario
            FROM Usuario u
            JOIN auth_user au ON LOWER(au.email) = LOWER(u.Correo)
            WHERE au.id = %s
        """, [user.id])
        row = cur.fetchone()

    if not row:
        return JsonResponse({"rows": []})

    id_aprendiz = row[0]

    # Buscar instructores asignados
    with connection.cursor() as cur:
        cur.execute("""
            SELECT ia.ID_Instructor,
                   COALESCE(
                     NULLIF(LTRIM(RTRIM(au.first_name + ' ' + au.last_name)), ''),
                     au.username
                   ) AS Nombre,
                   ia.TipoCompetencia,
                   ia.Trimestre
            FROM Instructor_Asignado ia
            JOIN Usuario u ON u.ID_Usuario = ia.ID_Instructor
            JOIN auth_user au ON LOWER(au.email) = LOWER(u.Correo)
            WHERE ia.ID_Aprendiz = %s
            ORDER BY ia.TipoCompetencia, Nombre
        """, [id_aprendiz])
        rows = cur.fetchall()

    out = []
    for iid, nombre, tipo, tri in rows:
        out.append({
            "id": iid,
            "nombre": nombre or "",
            "tipo": tipo or "",
            "trimestre": tri or "",
        })

    return JsonResponse({"rows": out})


@aprendiz_required
@require_POST
def justificacion_create_api(request):
    """
    Recibe multipart/form-data:
    - motivo (str)
    - fecha (YYYY-MM-DD)  => FechaInasistencia
    - observacion (str, opcional)
    - adjunto (file, opcional)
    - instructor (ID_Usuario del instructor destino, desde Instructor_Asignado)

    Inserta en [Justificaciones] y crea un registro en [Justificacion_Destino].
    Si no existe Asistencia para esa fecha del usuario, crea un registro 'Ausente'
    para poder referenciarlo.
    """
    user = request.user
    motivo = (request.POST.get("motivo") or "").strip()
    observacion = (request.POST.get("observacion") or "").strip()
    fecha_s = (request.POST.get("fecha") or "").strip()
    instructor_s = (request.POST.get("instructor") or "").strip()

    if not motivo or not fecha_s:
        return JsonResponse({"ok": False, "msg": "Completa motivo y fecha."}, status=400)

    try:
        fecha_inas = datetime.strptime(fecha_s, "%Y-%m-%d").date()
    except Exception:
        return JsonResponse({"ok": False, "msg": "Fecha inválida."}, status=400)

    try:
        id_instructor = int(instructor_s or 0)
    except Exception:
        id_instructor = 0

    if id_instructor <= 0:
        return JsonResponse({"ok": False, "msg": "Selecciona el instructor al que va la justificación."}, status=400)

    # Mapear a Usuario.ID_Usuario e ID_Programa (aprendiz)
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 u.ID_Usuario, u.ID_Programa
            FROM Usuario u
            JOIN auth_user au ON LOWER(au.email) = LOWER(u.Correo)
            WHERE au.id = %s
        """, [user.id])
        row = cur.fetchone()

    if not row:
        return JsonResponse({"ok": False, "msg": "No estás asociado en la tabla Usuario."}, status=403)

    usuario_id, id_programa = row

    # Validar que ese instructor esté asignado a este aprendiz
    with connection.cursor() as cur:
        cur.execute("""
            SELECT COUNT(*)
            FROM Instructor_Asignado
            WHERE ID_Aprendiz = %s AND ID_Instructor = %s
        """, [usuario_id, id_instructor])
        c_row = cur.fetchone()

    if not c_row or c_row[0] == 0:
        return JsonResponse({"ok": False, "msg": "Ese instructor no está asignado a tu ficha/trimestre."}, status=400)

    # ===============================
    # Buscar o crear asistencia del día
    # ===============================
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 ID_Asistencia, Estado
            FROM Asistencia
            WHERE ID_Usuario = %s AND FechaRegistro = %s
            ORDER BY ID_Asistencia DESC
        """, [user.id, fecha_inas])
        asis = cur.fetchone()

    if asis:
        id_asistencia, _old_estado = asis
    else:
        # Crear un registro 'Ausente' para poder referenciarlo
        with connection.cursor() as cur:
            cur.execute("""
                INSERT INTO Asistencia (
                    TipoRegistro, FechaRegistro, HoraRegistro,
                    Estado, ID_Usuario, ID_Programa, ID_Codigo
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, [
                "Entrada",
                fecha_inas,
                "00:00:00",
                "Ausente",
                user.id,
                id_programa,
                None,
            ])

        # Releer el ID recién creado SIN usar SCOPE_IDENTITY()
        with connection.cursor() as cur:
            cur.execute("""
                SELECT TOP 1 ID_Asistencia
                FROM Asistencia
                WHERE ID_Usuario = %s AND FechaRegistro = %s
                ORDER BY ID_Asistencia DESC
            """, [user.id, fecha_inas])
            row_as = cur.fetchone()

        if not row_as:
            return JsonResponse(
                {"ok": False, "msg": "No pude ubicar el registro de asistencia recién creado."},
                status=500,
            )

        id_asistencia = row_as[0]

    # ===============================
    # Guardar archivo (si viene)
    # ===============================
    adj_rel = None
    if "adjunto" in request.FILES:
        try:
            adj_rel = _save_uploaded_file(user.id, request.FILES["adjunto"])
        except Exception:
            adj_rel = None

    hoy = timezone.localtime().date()

    # Insertar Justificación y obtener el ID_Justificacion
    with connection.cursor() as cur:
        cur.execute("""
            INSERT INTO Justificaciones (
                FechaEnvio, ArchivoAdjunto, FechaInasistencia,
                Observacion, Estado, Motivo, ID_Usuario, ID_Asistencia
            )
            OUTPUT INSERTED.ID_Justificacion
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, [hoy, adj_rel, fecha_inas, observacion, "Pendiente", motivo, usuario_id, id_asistencia])
        jrow = cur.fetchone()

    id_justificacion = jrow[0]

    # Relación Justificación -> Instructor destino
    with connection.cursor() as cur:
        cur.execute("""
            INSERT INTO Justificacion_Destino (ID_Justificacion, ID_Instructor)
            VALUES (%s, %s)
        """, [id_justificacion, id_instructor])

    # Notificar SOLO al instructor elegido
    titulo = "Nueva justificación recibida"
    cuerpo = f"{user.get_full_name() or user.username} • {fecha_inas} • Motivo: {motivo}"

    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 au.id
            FROM Usuario u
            JOIN auth_user au ON LOWER(au.email) = LOWER(u.Correo)
            WHERE u.ID_Usuario = %s
        """, [id_instructor])
        row = cur.fetchone()

    if row:
        instructor_auth_id = row[0]
        with connection.cursor() as cur:
            cur.execute("""
                INSERT INTO Notificacion (ID_Usuario, Titulo, Cuerpo, Tipo)
                VALUES (%s, %s, %s, %s)
            """, [instructor_auth_id, titulo, cuerpo, "info"])

    return JsonResponse({"ok": True, "msg": "Justificación enviada."})


@aprendiz_required
@require_GET
def justificacion_list_api(request):
    """
    Lista de justificaciones del aprendiz actual, basada en tu tabla [Justificaciones].
    """
    user = request.user
    rows = []

    # Primero mapear al ID_Usuario de tu tabla Usuario
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 u.ID_Usuario
           	FROM Usuario u
            JOIN auth_user au ON LOWER(au.email)=LOWER(u.Correo)
            WHERE au.id=%s
        """, [user.id])
        link = cur.fetchone()

    if not link:
        return JsonResponse({"rows": rows})

    usuario_id = link[0]

    with connection.cursor() as cur:
        cur.execute("""
            SELECT ID_Justificacion, Motivo, Observacion, FechaEnvio, FechaInasistencia, ArchivoAdjunto, Estado, ID_Asistencia
            FROM Justificaciones
            WHERE ID_Usuario = %s
            ORDER BY ID_Justificacion DESC
        """, [usuario_id])
        for (jid, mot, obs, fenv, finas, adj, est, ida) in cur.fetchall():
            rows.append({
                "id": jid,
                "motivo": mot or "",
                "observacion": obs or "",
                "fecha_envio": fenv.isoformat() if fenv else "",
                "fecha_inasistencia": finas.isoformat() if finas else "",
                "adjunto": adj or "",
                "estado": est or "Pendiente",
                "id_asistencia": ida or 0,
            })

    return JsonResponse({"rows": rows})


# ---------- Instructor ----------

@instructor_required
@never_cache
def justificaciones_instructor(request):
    return render(request, "dash/justificaciones_instructor.html")


@instructor_required
@require_GET
def instructor_justif_list_api(request):
    """
    Lista de justificaciones para el instructor logueado.

    ?estado=pendiente|aprobado|rechazado|todas (default pendiente)
    ?q=texto  -> busca por motivo / observación / nombre de aprendiz
    """
    estado = (request.GET.get("estado") or "pendiente").strip().lower()
    term = (request.GET.get("q") or "").strip()

    # 1) Mapear instructor actual -> Usuario.ID_Usuario
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 u.ID_Usuario
            FROM Usuario u
            JOIN auth_user au ON LOWER(au.email) = LOWER(u.Correo)
            WHERE au.id = %s
        """, [request.user.id])
        r = cur.fetchone()

    if not r:
        return JsonResponse({"ok": True, "rows": []})

    id_instructor = r[0]

    # 2) Filtros
    where_clauses = ["jd.ID_Instructor = %s"]
    params = [id_instructor]

    if estado in ("pendiente", "aprobado", "rechazado"):
        where_clauses.append("j.Estado = %s")
        params.append(estado.capitalize())

    if term:
        like = f"%{term}%"
        where_clauses.append("""
            (
                j.Motivo LIKE %s
                OR j.Observacion LIKE %s
                OR au.first_name LIKE %s
                OR au.last_name LIKE %s
                OR au.username LIKE %s
            )
        """)
        params.extend([like, like, like, like, like])

    where_sql = "WHERE " + " AND ".join(where_clauses)

    rows = []
    with connection.cursor() as cur:
        cur.execute(f"""
            SELECT j.ID_Justificacion,
                   j.Motivo, j.Observacion, j.FechaEnvio, j.FechaInasistencia,
                   j.ArchivoAdjunto, j.Estado,
                   j.ID_Asistencia,
                   u.ID_Usuario,
                   COALESCE(
                     NULLIF(LTRIM(RTRIM(au.first_name + ' ' + au.last_name)), ''),
                     au.username
                   ) AS Nombre
            FROM Justificaciones j
            JOIN Justificacion_Destino jd ON jd.ID_Justificacion = j.ID_Justificacion
            JOIN Usuario u ON u.ID_Usuario = j.ID_Usuario
            JOIN auth_user au ON LOWER(au.email) = LOWER(u.Correo)
            {where_sql}
            ORDER BY j.ID_Justificacion DESC
        """, params)
        for (jid, mot, obs, fenv, finas, adj, est, ida, uidu, nombre) in cur.fetchall():
            rows.append({
                "id": jid,
                "motivo": mot or "",
                "observacion": obs or "",
                "fecha_envio": fenv.isoformat() if fenv else "",
                "fecha_inasistencia": finas.isoformat() if finas else "",
                "adjunto": adj or "",
                "estado": est or "Pendiente",
                "id_asistencia": ida or 0,
                "usuario_id": uidu,
                "nombre": nombre or "",
            })
    return JsonResponse({"ok": True, "rows": rows})


@instructor_required
@require_POST
def instructor_justif_set_state_api(request):
    """
    JSON:
    { "id": <id_justificacion>, "accion": "aprobar"|"rechazar", "observacion": "..." }
    Cambia estado en [Justificaciones]. Si apruebas:
      - Se pone Asistencia.Estado = 'Justificado' en el ID_Asistencia referenciado.
      - Se inserta en Historial_Asistencia.
      - Se notifica al aprendiz.
    """
    import json
    body = json.loads(request.body.decode("utf-8") or "{}")
    try:
        jid = int(body.get("id") or 0)
    except Exception:
        jid = 0
    accion = (body.get("accion") or "").strip().lower()
    obs = (body.get("observacion") or "").strip()

    if jid <= 0 or accion not in ("aprobar", "rechazar"):
        return JsonResponse({"ok": False, "msg": "Datos inválidos."}, status=400)

    # Leer datos base
    with connection.cursor() as cur:
        cur.execute("""
            SELECT j.ID_Asistencia, j.Estado, j.ID_Usuario, j.FechaInasistencia
            FROM Justificaciones j
            WHERE j.ID_Justificacion = %s
        """, [jid])
        row = cur.fetchone()

    if not row:
        return JsonResponse({"ok": False, "msg": "Justificación no existe."}, status=404)

    id_asistencia, estado_actual, id_usuario_tbl, fecha_inas = row
    if estado_actual in ("Aprobado", "Rechazado"):
        return JsonResponse({"ok": False, "msg": "Esta solicitud ya fue resuelta."}, status=400)

    nuevo_estado = "Aprobado" if accion == "aprobar" else "Rechazado"

    # Mapear ID_Usuario (tabla Usuario) -> auth_user.id
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 au.id
            FROM Usuario u
            JOIN auth_user au ON LOWER(au.email)=LOWER(u.Correo)
            WHERE u.ID_Usuario=%s
        """, [id_usuario_tbl])
        aurow = cur.fetchone()
    if not aurow:
        return JsonResponse({"ok": False, "msg": "No pude mapear el usuario."}, status=500)
    auth_uid = aurow[0]

    # Update estado y observación
    with connection.cursor() as cur:
        cur.execute("""
            UPDATE Justificaciones
               SET Estado = %s,
                   Observacion = %s
             WHERE ID_Justificacion = %s
        """, [nuevo_estado, obs, jid])

    # Notificar al aprendiz
    titulo = f"Justificación {nuevo_estado.lower()}"
    cuerpo = f"Estado: {nuevo_estado}. Fecha: {fecha_inas}."
    with connection.cursor() as cur:
        cur.execute("""
            INSERT INTO Notificacion (ID_Usuario, Titulo, Cuerpo, Tipo)
            VALUES (%s, %s, %s, %s)
        """, [auth_uid, titulo, cuerpo, "info" if nuevo_estado == "Aprobado" else "warn"])

    # Si apruebas: cambiar Asistencia y log en Historial_Asistencia
    if nuevo_estado == "Aprobado" and id_asistencia:
        try:
            # Estado anterior
            with connection.cursor() as cur:
                cur.execute("""
                    SELECT TOP 1 Estado
                    FROM Asistencia
                    WHERE ID_Asistencia = %s
                """, [id_asistencia])
                row2 = cur.fetchone()
            if row2:
                old_state = row2[0] or ""

                with connection.cursor() as cur:
                    cur.execute("""
                        UPDATE Asistencia
                           SET Estado = 'Justificado'
                         WHERE ID_Asistencia = %s
                    """, [id_asistencia])

                with connection.cursor() as cur:
                    cur.execute("""
                        INSERT INTO Historial_Asistencia
                          (ID_Asistencia, ID_Usuario, CampoModificado, ValorAnterior, ValorNuevo, FechaCambio)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, [id_asistencia, auth_uid, 'Estado', old_state, 'Justificado', date.today()])
        except Exception:
            # no se cae todo si el historial falla
            pass

    return JsonResponse({"ok": True, "msg": f"Estado actualizado a {nuevo_estado}."})


# ============================================================
# ================== REPORTES INSTRUCTOR ======================
# ============================================================

@instructor_required
@never_cache
def instructor_reportes(request):
    """
    Reportes de asistencia para el instructor.
    Usa Asistencia + Instructor_Asignado + Programa.
    Soporta:
      - Filtro por rango de días (?dias=7,15,30,...)
      - Filtro por programa/ficha (?programa_id=ID_Programa)
      - Detalle por aprendiz cuando se envía programa_id
    """
    # Rango de días
    try:
        dias = int(request.GET.get("dias", "30"))
    except Exception:
        dias = 30
    dias = max(1, min(dias, 365))

    hoy = date.today()
    desde = hoy - timedelta(days=dias)

    # Filtro opcional por programa
    programa_raw = (request.GET.get("programa_id") or "").strip()
    try:
        programa_id = int(programa_raw)
        if programa_id <= 0:
            programa_id = None
    except Exception:
        programa_id = None

    # Mapear instructor actual -> Usuario.ID_Usuario
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 u.ID_Usuario
            FROM Usuario u
            JOIN auth_user au ON LOWER(au.email) = LOWER(u.Correo)
            WHERE au.id = %s
        """, [request.user.id])
        r = cur.fetchone()

    if not r:
        context = {
            "sin_instructor": True,
            "sin_aprendices": False,
            "dias": dias,
            "desde": desde,
            "hasta": hoy,
        }
        return render(request, "dash/instructor_reportes.html", context)

    id_instructor = r[0]

    # ¿Tiene aprendices asignados?
    with connection.cursor() as cur:
        cur.execute("""
            SELECT COUNT(*)
            FROM Instructor_Asignado
            WHERE ID_Instructor = %s
        """, [id_instructor])
        row_asig = cur.fetchone()

    asignados = row_asig[0] if row_asig else 0
    if asignados == 0:
        context = {
            "sin_instructor": False,
            "sin_aprendices": True,
            "dias": dias,
            "desde": desde,
            "hasta": hoy,
            "total": 0,
        }
        return render(request, "dash/instructor_reportes.html", context)

    # =========================
    # Resumen global de asistencia
    # =========================
    filtros_sql = """
        FROM Asistencia a
        JOIN auth_user au ON au.id = a.ID_Usuario
        JOIN Usuario u ON LOWER(au.email) = LOWER(u.Correo)
        JOIN Instructor_Asignado ia ON ia.ID_Aprendiz = u.ID_Usuario
        WHERE ia.ID_Instructor = %s
          AND a.FechaRegistro BETWEEN %s AND %s
    """
    filtros_params = [id_instructor, desde, hoy]

    if programa_id is not None:
        filtros_sql += " AND a.ID_Programa = %s"
        filtros_params.append(programa_id)

    with connection.cursor() as cur:
        cur.execute(f"""
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN a.Estado IN ('Presente','Justificado') THEN 1 ELSE 0 END) AS presentes,
                SUM(CASE WHEN a.Estado = 'Tarde' THEN 1 ELSE 0 END) AS tardes,
                SUM(CASE WHEN a.Estado = 'Ausente' THEN 1 ELSE 0 END) AS ausentes
            {filtros_sql}
        """, filtros_params)
        row = cur.fetchone()

    total = row[0] if row and row[0] is not None else 0
    presentes = row[1] if row and row[1] is not None else 0
    tardes = row[2] if row and row[2] is not None else 0
    ausentes = row[3] if row and row[3] is not None else 0

    if total:
        pct_presente = round(presentes * 100.0 / total, 1)
        pct_tarde = round(tardes * 100.0 / total, 1)
        pct_ausente = round(ausentes * 100.0 / total, 1)
    else:
        pct_presente = pct_tarde = pct_ausente = 0.0

    # =========================
    # Detalle por programa / ficha (tabla de arriba)
    # =========================
    with connection.cursor() as cur:
        sql_por_programa = """
            SELECT
                p.ID_Programa,
                p.NombrePrograma,
                p.Ficha,
                COUNT(*) AS total,
                SUM(CASE WHEN a.Estado IN ('Presente','Justificado') THEN 1 ELSE 0 END) AS presentes,
                SUM(CASE WHEN a.Estado = 'Tarde' THEN 1 ELSE 0 END) AS tardes,
                SUM(CASE WHEN a.Estado = 'Ausente' THEN 1 ELSE 0 END) AS ausentes
            FROM Asistencia a
            JOIN Programa p ON p.ID_Programa = a.ID_Programa
            JOIN auth_user au ON au.id = a.ID_Usuario
            JOIN Usuario u ON LOWER(au.email) = LOWER(u.Correo)
            JOIN Instructor_Asignado ia ON ia.ID_Aprendiz = u.ID_Usuario
            WHERE ia.ID_Instructor = %s
              AND a.FechaRegistro BETWEEN %s AND %s
        """
        params_prog = [id_instructor, desde, hoy]

        if programa_id is not None:
            sql_por_programa += " AND a.ID_Programa = %s"
            params_prog.append(programa_id)

        sql_por_programa += """
            GROUP BY p.ID_Programa, p.NombrePrograma, p.Ficha
            ORDER BY p.NombrePrograma, p.Ficha
        """

        cur.execute(sql_por_programa, params_prog)
        rows = cur.fetchall()

    por_programa = []
    for (id_prog, nombre_prog, ficha, tot, pres, tar, aus) in rows:
        tot = tot or 0
        pres = pres or 0
        tar = tar or 0
        aus = aus or 0
        if tot:
            pct = round(pres * 100.0 / tot, 1)
        else:
            pct = 0.0
        por_programa.append({
            "programa_id": id_prog,
            "nombre_programa": nombre_prog or "",
            "ficha": ficha or "",
            "total": tot,
            "presentes": pres,
            "tardes": tar,
            "ausentes": aus,
            "pct_presente": pct,
        })

    # =========================
    # Datos para la gráfica
    # =========================
    import json
    labels = [p["ficha"] or "" for p in por_programa]
    data_presentes = [p["presentes"] for p in por_programa]
    data_ausentes = [p["ausentes"] for p in por_programa]

    chart_json = json.dumps({
        "labels": labels,
        "presentes": data_presentes,
        "ausentes": data_ausentes,
    })

    # =========================
    # Detalle por aprendices en una ficha (panel de abajo)
    # =========================
    ficha_meta = None
    detalle_ficha = []

    if programa_id is not None:
        # Datos de la ficha
        with connection.cursor() as cur:
            cur.execute("""
                SELECT TOP 1 NombrePrograma, Ficha, Jornada
                FROM Programa
                WHERE ID_Programa = %s
            """, [programa_id])
            fm = cur.fetchone()

        if fm:
            ficha_meta = {
                "nombre_programa": fm[0] or "",
                "ficha": fm[1] or "",
                "jornada": fm[2] or "",
            }

            # Detalle por aprendiz
            with connection.cursor() as cur:
                cur.execute("""
                    SELECT
                        COALESCE(
                          NULLIF(LTRIM(RTRIM(au.first_name + ' ' + au.last_name)), ''),
                          au.username
                        ) AS Nombre,
                        au.email,
                        COUNT(*) AS total,
                        SUM(CASE WHEN a.Estado IN ('Presente','Justificado') THEN 1 ELSE 0 END) AS presentes,
                        SUM(CASE WHEN a.Estado = 'Tarde' THEN 1 ELSE 0 END) AS tardes,
                        SUM(CASE WHEN a.Estado = 'Ausente' THEN 1 ELSE 0 END) AS ausentes,
                        SUM(CASE WHEN a.Estado = 'Justificado' THEN 1 ELSE 0 END) AS justificados
                    FROM Asistencia a
                    JOIN auth_user au ON au.id = a.ID_Usuario
                    JOIN Usuario u ON LOWER(au.email) = LOWER(u.Correo)
                    JOIN Instructor_Asignado ia ON ia.ID_Aprendiz = u.ID_Usuario
                    WHERE ia.ID_Instructor = %s
                      AND a.ID_Programa = %s
                      AND a.FechaRegistro BETWEEN %s AND %s
                    GROUP BY
                        COALESCE(
                          NULLIF(LTRIM(RTRIM(au.first_name + ' ' + au.last_name)), ''),
                          au.username
                        ),
                        au.email
                    ORDER BY Nombre
                """, [id_instructor, programa_id, desde, hoy])
                rows_det = cur.fetchall()

            for (nombre, email, tot, pres, tar, aus, justif) in rows_det:
                tot = tot or 0
                pres = pres or 0
                tar = tar or 0
                aus = aus or 0
                justif = justif or 0
                pct = round(pres * 100.0 / tot, 1) if tot else 0.0
                detalle_ficha.append({
                    "nombre": nombre or "",
                    "email": email or "",
                    "total": tot,
                    "presentes": pres,
                    "tardes": tar,
                    "ausentes": aus,
                    "justificados": justif,
                    "pct_presente": pct,
                })

    context = {
        "sin_instructor": False,
        "sin_aprendices": False,
        "dias": dias,
        "desde": desde,
        "hasta": hoy,
        "total": total,
        "presentes": presentes,
        "tardes": tardes,
        "ausentes": ausentes,
        "pct_presente": pct_presente,
        "pct_tarde": pct_tarde,
        "pct_ausente": pct_ausente,
        "por_programa": por_programa,
        "chart_json": chart_json,
        "programa_id": programa_id,
        "ficha_meta": ficha_meta,
        "detalle_ficha": detalle_ficha,
    }
    return render(request, "dash/instructor_reportes.html", context)


# ============================================================
# ======== HISTORIAL DE ASISTENCIA (INSTRUCTOR NUEVO) ========
# ============================================================

@instructor_required
@never_cache
def instructor_historial_view(request):
    """
    Pantalla del historial de asistencia del instructor
    (ve asistencias de sus aprendices).
    """
    return render(request, "dash/instructor_historial.html")


@instructor_required
@require_GET
def instructor_historial_api(request):
    """
    Devuelve el historial de asistencia de los aprendices del instructor logueado.

    Filtros:
      - dias (int, por defecto 30)
      - estado (Presente, Tarde, Ausente, Justificado, 'todos')
      - q (busca por nombre, correo, programa, ficha)
      - page, page_size
    """
    user = request.user

    # Mapear instructor actual -> Usuario.ID_Usuario
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 u.ID_Usuario
            FROM Usuario u
            JOIN auth_user au ON LOWER(au.email) = LOWER(u.Correo)
            WHERE au.id = %s
        """, [user.id])
        row = cur.fetchone()

    if not row:
        return JsonResponse({
            "rows": [],
            "page": 1,
            "pages": 1,
            "total": 0,
            "stats": {
                "total": 0,
                "presentes": 0,
                "tardes": 0,
                "ausentes": 0,
                "justificados": 0,
            }
        })

    id_instructor = row[0]

    # Parámetros
    try:
        dias = int(request.GET.get("dias", "30"))
    except Exception:
        dias = 30
    dias = max(1, min(dias, 365))

    hoy = date.today()
    desde = hoy - timedelta(days=dias)

    estado = (request.GET.get("estado") or "").strip()
    q = (request.GET.get("q") or "").strip()

    try:
        page = max(int(request.GET.get("page", "1")), 1)
    except Exception:
        page = 1
    try:
        page_size = int(request.GET.get("page_size", "12"))
    except Exception:
        page_size = 12
    page_size = max(1, min(page_size, 50))

    where_clauses = [
        "ia.ID_Instructor = %s",
        "a.FechaRegistro BETWEEN %s AND %s",
    ]
    params = [id_instructor, desde, hoy]

    if estado and estado.lower() != "todos":
        where_clauses.append("a.Estado = %s")
        params.append(estado)

    if q:
        like = f"%{q}%"
        where_clauses.append("""
            (
                au.first_name LIKE %s
                OR au.last_name LIKE %s
                OR au.username LIKE %s
                OR au.email LIKE %s
                OR p.NombrePrograma LIKE %s
                OR p.Ficha LIKE %s
            )
        """)
        params.extend([like, like, like, like, like, like])

    where_sql = "WHERE " + " AND ".join(where_clauses)

    # Conteo y stats
    with connection.cursor() as cur:
        cur.execute(f"""
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN a.Estado IN ('Presente','Justificado') THEN 1 ELSE 0 END) AS presentes,
                SUM(CASE WHEN a.Estado = 'Tarde' THEN 1 ELSE 0 END) AS tardes,
                SUM(CASE WHEN a.Estado = 'Ausente' THEN 1 ELSE 0 END) AS ausentes,
                SUM(CASE WHEN a.Estado = 'Justificado' THEN 1 ELSE 0 END) AS justificados
            FROM Asistencia a
            JOIN auth_user au ON au.id = a.ID_Usuario
            JOIN Usuario u ON LOWER(au.email) = LOWER(u.Correo)
            JOIN Instructor_Asignado ia ON ia.ID_Aprendiz = u.ID_Usuario
            JOIN Programa p ON p.ID_Programa = a.ID_Programa
            {where_sql}
        """, params)
        row_stats = cur.fetchone()

    total = row_stats[0] if row_stats and row_stats[0] is not None else 0
    presentes = row_stats[1] if row_stats and row_stats[1] is not None else 0
    tardes = row_stats[2] if row_stats and row_stats[2] is not None else 0
    ausentes = row_stats[3] if row_stats and row_stats[3] is not None else 0
    justificados = row_stats[4] if row_stats and row_stats[4] is not None else 0

    pages = max((total + page_size - 1) // page_size, 1)
    if page > pages:
        page = pages
    offset = (page - 1) * page_size

    rows_out = []
    if total > 0:
        with connection.cursor() as cur:
            cur.execute(f"""
                SELECT
                    a.ID_Asistencia,
                    a.TipoRegistro,
                    a.FechaRegistro,
                    a.HoraRegistro,
                    a.Estado,
                    p.NombrePrograma,
                    p.Ficha,
                    p.Jornada,
                    COALESCE(
                      NULLIF(LTRIM(RTRIM(au.first_name + ' ' + au.last_name)), ''),
                      au.username
                    ) AS Nombre,
                    au.email
                FROM Asistencia a
                JOIN auth_user au ON au.id = a.ID_Usuario
                JOIN Usuario u ON LOWER(au.email) = LOWER(u.Correo)
                JOIN Instructor_Asignado ia ON ia.ID_Aprendiz = u.ID_Usuario
                JOIN Programa p ON p.ID_Programa = a.ID_Programa
                {where_sql}
                ORDER BY a.FechaRegistro DESC, a.HoraRegistro DESC, a.ID_Asistencia DESC
                OFFSET %s ROWS FETCH NEXT %s ROWS ONLY
            """, params + [offset, page_size])
            for (ida, tipo, f, h, estado_row, prog, ficha, jornada, nombre, email) in cur.fetchall():
                rows_out.append({
                    "id": ida,
                    "tipo": tipo or "",
                    "fecha": f.isoformat() if f else "",
                    "hora": h.strftime("%H:%M:%S") if h else "",
                    "estado": estado_row or "",
                    "programa": prog or "",
                    "ficha": ficha or "",
                    "jornada": jornada or "",
                    "aprendiz": nombre or "",
                    "email": email or "",
                })

    return JsonResponse({
        "rows": rows_out,
        "page": page,
        "pages": pages,
        "total": total,
        "stats": {
            "total": total,
            "presentes": presentes,
            "tardes": tardes,
            "ausentes": ausentes,
            "justificados": justificados,
        }
    })


# ============================================================
# ========= EXPORTAR HISTORIAL INSTRUCTOR A EXCEL =============
# ============================================================

@instructor_required
def instructor_historial_excel(request):
    """
    Exporta a Excel el mismo historial que ve el instructor,
    respetando filtros: dias, estado, q.
    """
    from io import BytesIO
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return HttpResponse(
            "Falta instalar openpyxl (pip install openpyxl) para exportar a Excel.",
            content_type="text/plain",
            status=500,
        )

    user = request.user

    # Mapear instructor actual -> Usuario.ID_Usuario
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 u.ID_Usuario
            FROM Usuario u
            JOIN auth_user au ON LOWER(au.email) = LOWER(u.Correo)
            WHERE au.id = %s
        """, [user.id])
        row = cur.fetchone()

    if not row:
        return HttpResponse(
            "Instructor no mapeado en la tabla Usuario.",
            content_type="text/plain",
            status=404,
        )

    id_instructor = row[0]

    # Filtros
    try:
        dias = int(request.GET.get("dias", "30"))
    except Exception:
        dias = 30
    dias = max(1, min(dias, 365))

    hoy = date.today()
    desde = hoy - timedelta(days=dias)

    estado = (request.GET.get("estado") or "").strip()
    q = (request.GET.get("q") or "").strip()

    where_clauses = [
        "ia.ID_Instructor = %s",
        "a.FechaRegistro BETWEEN %s AND %s",
    ]
    params = [id_instructor, desde, hoy]

    if estado and estado.lower() != "todos":
        where_clauses.append("a.Estado = %s")
        params.append(estado)

    if q:
        like = f"%{q}%"
        where_clauses.append("""
            (
                au.first_name LIKE %s
                OR au.last_name LIKE %s
                OR au.username LIKE %s
                OR au.email LIKE %s
                OR p.NombrePrograma LIKE %s
                OR p.Ficha LIKE %s
            )
        """)
        params.extend([like, like, like, like, like, like])

    where_sql = "WHERE " + " AND ".join(where_clauses)

    # Traer todas las filas
    with connection.cursor() as cur:
        cur.execute(f"""
            SELECT
                a.ID_Asistencia,
                a.TipoRegistro,
                a.FechaRegistro,
                a.HoraRegistro,
                a.Estado,
                p.NombrePrograma,
                p.Ficha,
                p.Jornada,
                COALESCE(
                  NULLIF(LTRIM(RTRIM(au.first_name + ' ' + au.last_name)), ''),
                  au.username
                ) AS Nombre,
                au.email
            FROM Asistencia a
            JOIN auth_user au ON au.id = a.ID_Usuario
            JOIN Usuario u ON LOWER(au.email) = LOWER(u.Correo)
            JOIN Instructor_Asignado ia ON ia.ID_Aprendiz = u.ID_Usuario
            JOIN Programa p ON p.ID_Programa = a.ID_Programa
            {where_sql}
            ORDER BY a.FechaRegistro DESC, a.HoraRegistro DESC, a.ID_Asistencia DESC
        """, params)
        rows = cur.fetchall()

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Encabezado descriptivo
    ws["A1"] = "Historial de asistencia"
    ws["A1"].font = title_font
    ws.merge_cells("A1:J1")

    ws["A2"] = f"Instructor: {user.get_full_name() or user.username}"
    ws.merge_cells("A2:J2")

    generado_txt = timezone.localtime().strftime("%Y-%m-%d %H:%M")
    ws["A3"] = f"Generado: {generado_txt}"
    ws.merge_cells("A3:J3")

    ws["A4"] = f"Rango: {desde} a {hoy} (últimos {dias} días)"
    ws.merge_cells("A4:J4")

    ws["A5"] = f"Estado: {estado or 'todos'}"
    ws.merge_cells("A5:J5")

    ws["A6"] = f"Búsqueda: {q or '—'}"
    ws.merge_cells("A6:J6")

    start_row = 8
    headers = [
        "Fecha",
        "Hora",
        "Estado",
        "Tipo registro",
        "Nombre aprendiz",
        "Correo",
        "Programa",
        "Ficha",
        "Jornada",
        "ID asistencia",
    ]

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Datos
    row_idx = start_row + 1
    for (ida, tipo, f, h, estado_row, prog, ficha, jornada, nombre, email) in rows:
        values = [
            f.strftime("%Y-%m-%d") if f else "",
            h.strftime("%H:%M:%S") if h else "",
            estado_row or "",
            tipo or "",
            nombre or "",
            email or "",
            prog or "",
            ficha or "",
            jornada or "",
            ida,
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = thin_border
        row_idx += 1

    # Ancho de columnas
    widths = [12, 10, 12, 16, 26, 26, 22, 10, 12, 14]
    for idx, w in enumerate(widths, start=1):
        col_letter = chr(64 + idx)  # A,B,C...
        ws.column_dimensions[col_letter].width = w

    # Congelar encabezado
    ws.freeze_panes = f"A{start_row+1}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"historial_asistencia_{hoy.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ============================================================
# ======= ALERTAS DE DESERCIÓN TEMPRANA (INSTRUCTOR) =========
# ============================================================

@instructor_required
@never_cache
def instructor_alertas_desercion_view(request):
    """
    Pantalla de alertas de posible deserción temprana.
    Los datos se cargan vía JS desde instructor_alertas_desercion_api.
    """
    default_dias = 30
    default_min_ausencias = 3
    default_max_pct_asistencia = 70
    context = {
        "default_dias": default_dias,
        "default_min_ausencias": default_min_ausencias,
        "default_max_pct_asistencia": default_max_pct_asistencia,
    }
    return render(request, "dash/instructor_alertas_desercion.html", context)


@instructor_required
@require_GET
def instructor_alertas_desercion_api(request):
    """
    Devuelve lista de aprendices "en riesgo" para el instructor logueado.

    Filtros GET:
      - dias (int, por defecto 30)
      - riesgo (Alto|Medio|Bajo|todos)
      - programa_id (opcional)
      - q (texto para buscar por nombre/correo/programa/ficha)

    Lógica de riesgo (según número de inasistencias EN EL RANGO DE DÍAS):
      - 3 o más inasistencias  -> Alto
      - 2 inasistencias        -> Medio
      - 1 inasistencia         -> Bajo
      - 0 inasistencias        -> no aparece

    La API devuelve, por cada aprendiz:
      - nombre, email
      - programa, ficha, jornada
      - total_sesiones
      - inasistencias
      - pct_asistencia (0–100)
      - ultima_asistencia (YYYY-MM-DD o null)
      - dias_sin_asistir (int o null)
      - riesgo ("Alto" | "Medio" | "Bajo")

    y en stats:
      - total  -> cuántos aprendices aparecen en la tabla
      - alto   -> cuántos en riesgo alto
      - medio  -> cuántos en riesgo medio
      - bajo   -> cuántos en riesgo bajo
      - dias   -> rango usado
    """
    user = request.user

    # Mapear instructor actual -> Usuario.ID_Usuario
    with connection.cursor() as cur:
        cur.execute("""
            SELECT TOP 1 u.ID_Usuario
            FROM Usuario u
            JOIN auth_user au ON LOWER(au.email) = LOWER(u.Correo)
            WHERE au.id = %s
        """, [user.id])
        row = cur.fetchone()

    if not row:
        # Instructor no mapeado
        return JsonResponse({
            "rows": [],
            "stats": {
                "total": 0,
                "alto": 0,
                "medio": 0,
                "bajo": 0,
                "dias": 0,
            }
        })

    id_instructor = row[0]

    # ====== Filtros básicos ======
    try:
        dias = int(request.GET.get("dias", "30"))
    except Exception:
        dias = 30
    dias = max(1, min(dias, 365))

    hoy = date.today()
    desde = hoy - timedelta(days=dias)

    riesgo_filter = (request.GET.get("riesgo") or "").strip().lower()
    q = (request.GET.get("q") or "").strip()

    programa_raw = (request.GET.get("programa_id") or "").strip()
    try:
        programa_id = int(programa_raw)
        if programa_id <= 0:
            programa_id = None
    except Exception:
        programa_id = None

    where_clauses = [
        "ia.ID_Instructor = %s",
        "a.FechaRegistro BETWEEN %s AND %s",
    ]
    params = [id_instructor, desde, hoy]

    if programa_id is not None:
        where_clauses.append("a.ID_Programa = %s")
        params.append(programa_id)

    if q:
        like = f"%{q}%"
        where_clauses.append("""
            (
                au.first_name LIKE %s
                OR au.last_name LIKE %s
                OR au.username LIKE %s
                OR au.email LIKE %s
                OR p.NombrePrograma LIKE %s
                OR p.Ficha LIKE %s
            )
        """)
        params.extend([like, like, like, like, like, like])

    where_sql = "WHERE " + " AND ".join(where_clauses)

    # ====== Agregado por aprendiz (en el rango de días) ======
    with connection.cursor() as cur:
        cur.execute(f"""
            SELECT
                COALESCE(
                  NULLIF(LTRIM(RTRIM(au.first_name + ' ' + au.last_name)), ''),
                  au.username
                ) AS Nombre,
                au.email,
                p.NombrePrograma,
                p.Ficha,
                p.Jornada,
                COUNT(*) AS total_sesiones,
                SUM(CASE WHEN a.Estado = 'Ausente' THEN 1 ELSE 0 END) AS inasistencias,
                SUM(CASE WHEN a.Estado IN ('Presente','Justificado') THEN 1 ELSE 0 END) AS presentes,
                MAX(CASE
                      WHEN a.Estado IN ('Presente','Justificado')
                      THEN a.FechaRegistro
                    END) AS ultima_asistencia
            FROM Asistencia a
            JOIN auth_user au ON au.id = a.ID_Usuario
            JOIN Usuario u ON LOWER(au.email) = LOWER(u.Correo)
            JOIN Instructor_Asignado ia ON ia.ID_Aprendiz = u.ID_Usuario
            JOIN Programa p ON p.ID_Programa = a.ID_Programa
            {where_sql}
            GROUP BY
                COALESCE(
                  NULLIF(LTRIM(RTRIM(au.first_name + ' ' + au.last_name)), ''),
                  au.username
                ),
                au.email,
                p.NombrePrograma,
                p.Ficha,
                p.Jornada
            ORDER BY Nombre
        """, params)
        rows = cur.fetchall()

    out = []
    total_monitoreo = 0
    alto_count = 0
    medio_count = 0
    bajo_count = 0

    for (nombre, email, prog, ficha, jornada,
         total_sesiones, inasistencias, presentes, ultima_asistencia) in rows:

        total_sesiones = total_sesiones or 0
        inasistencias = inasistencias or 0
        presentes = presentes or 0

        # Porcentaje de asistencia
        if total_sesiones > 0:
            pct_asistencia = round(presentes * 100.0 / float(total_sesiones))
        else:
            pct_asistencia = 0.0

        # Días sin asistir (desde la última asistencia "real")
        if ultima_asistencia:
            try:
                dias_sin_asistir = (hoy - ultima_asistencia).days
                if dias_sin_asistir < 0:
                    dias_sin_asistir = 0
            except Exception:
                dias_sin_asistir = None
        else:
            dias_sin_asistir = None

        # ===== Lógica de riesgo por número de inasistencias =====
        if inasistencias >= 3:
            riesgo = "Alto"
        elif inasistencias == 2:
            riesgo = "Medio"
        elif inasistencias == 1:
            riesgo = "Bajo"
        else:
            riesgo = None

        # Si no tiene ninguna falta en el rango, no entra en monitoreo
        if not riesgo:
            continue

        # Filtro por riesgo desde el front (select)
        if riesgo_filter in ("alto", "medio", "bajo") and riesgo.lower() != riesgo_filter:
            continue

        # Contadores para los KPIs
        total_monitoreo += 1
        if riesgo == "Alto":
            alto_count += 1
        elif riesgo == "Medio":
            medio_count += 1
        elif riesgo == "Bajo":
            bajo_count += 1

        out.append({
            "nombre": nombre or "",
            "email": email or "",
            "programa": prog or "",
            "ficha": ficha or "",
            "jornada": jornada or "",
            "total_sesiones": int(total_sesiones),
            "inasistencias": int(inasistencias),
            "pct_asistencia": pct_asistencia,
            "ultima_asistencia": ultima_asistencia.isoformat() if ultima_asistencia else None,
            "dias_sin_asistir": dias_sin_asistir,
            "riesgo": riesgo,
        })

    return JsonResponse({
        "rows": out,
        "stats": {
            "total": total_monitoreo,
            "alto": alto_count,
            "medio": medio_count,
            "bajo": bajo_count,
            "dias": dias,
        }
    })
